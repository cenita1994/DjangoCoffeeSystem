from django.shortcuts import render
from django.utils.http import urlencode
from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from audittrail.utils import log_audit
from accounts.decorators import manager_or_owner_required
from .forms import ForecastFilterForm
from .services import (
    get_default_forecast_dates,
    build_ingredient_demand_forecast,
)


def get_forecast_filters(request):
    default_date_from, default_date_to = get_default_forecast_dates()

    form = ForecastFilterForm(request.GET or None)

    date_from = default_date_from
    date_to = default_date_to
    forecast_days = 7

    if form.is_valid():
        date_from = form.cleaned_data['date_from']
        date_to = form.cleaned_data['date_to']
        forecast_days = int(form.cleaned_data['forecast_days'])

        if date_to < date_from:
            date_from, date_to = date_to, date_from
    else:
        form = ForecastFilterForm(initial={
            'date_from': default_date_from,
            'date_to': default_date_to,
            'forecast_days': 7,
        })

    return form, date_from, date_to, forecast_days


@manager_or_owner_required
def forecasting_dashboard(request):
    form, date_from, date_to, forecast_days = get_forecast_filters(request)

    product_rows, detail_rows, ingredient_rows, product_summary, ingredient_summary = build_ingredient_demand_forecast(
        date_from=date_from,
        date_to=date_to,
        forecast_days=forecast_days
    )

    query_params = urlencode({
        'date_from': date_from.strftime('%Y-%m-%d'),
        'date_to': date_to.strftime('%Y-%m-%d'),
        'forecast_days': forecast_days,
    })

    log_audit(
        request=request,
        action='View',
        module='Demand Forecasting',
        description=(
            f'Generated Demand Forecasting dashboard. '
            f'Historical range: {date_from} to {date_to}. '
            f'Forecast horizon: next {forecast_days} days.'
        ),
        object_type='Forecast',
        object_repr='Demand Forecasting Dashboard'
    )

    return render(request, 'forecasting/dashboard.html', {
        'form': form,
        'date_from': date_from,
        'date_to': date_to,
        'forecast_days': forecast_days,
        'product_rows': product_rows,
        'detail_rows': detail_rows,
        'ingredient_rows': ingredient_rows,
        'product_summary': product_summary,
        'ingredient_summary': ingredient_summary,
        'query_params': query_params,
    })


@manager_or_owner_required
def export_forecasting_report(request):
    form, date_from, date_to, forecast_days = get_forecast_filters(request)

    product_rows, detail_rows, ingredient_rows, product_summary, ingredient_summary = build_ingredient_demand_forecast(
        date_from=date_from,
        date_to=date_to,
        forecast_days=forecast_days
    )

    log_audit(
        request=request,
        action='Export',
        module='Demand Forecasting',
        description=(
            f'Exported Demand Forecasting report. '
            f'Historical range: {date_from} to {date_to}. '
            f'Forecast horizon: next {forecast_days} days.'
        ),
        object_type='Forecast',
        object_repr='Demand Forecasting Excel Export'
    )

    workbook = Workbook()

    dark_fill = PatternFill(fill_type='solid', fgColor='212529')
    white_font = Font(color='FFFFFF', bold=True)
    bold_font = Font(bold=True)

    # Sheet 1: Summary
    summary_sheet = workbook.active
    summary_sheet.title = 'Forecast Summary'

    summary_sheet.append(['Django Coffee System'])
    summary_sheet.append(['Demand Forecasting Report'])
    summary_sheet.append(['Historical Date From', date_from.strftime('%Y-%m-%d')])
    summary_sheet.append(['Historical Date To', date_to.strftime('%Y-%m-%d')])
    summary_sheet.append(['Forecast Horizon', f'Next {forecast_days} days'])
    summary_sheet.append([])
    summary_sheet.append(['Metric', 'Value'])
    summary_sheet.append(['Products Forecasted', product_summary['product_count']])
    summary_sheet.append(['Historical Quantity Sold', float(product_summary['total_quantity_sold'])])
    summary_sheet.append(['Projected Quantity', float(product_summary['total_forecast_quantity'])])
    summary_sheet.append(['Projected Sales', float(product_summary['total_projected_sales'])])
    summary_sheet.append(['Projected Cost', float(product_summary['total_projected_cost'])])
    summary_sheet.append(['Projected Profit', float(product_summary['total_projected_profit'])])
    summary_sheet.append(['Ingredients Included', ingredient_summary['ingredient_count']])
    summary_sheet.append(['Ingredients with Restock Suggested', ingredient_summary['total_suggested_purchase_items']])

    for cell in summary_sheet[7]:
        cell.fill = dark_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal='center')

    # Sheet 2: Product Forecast
    product_sheet = workbook.create_sheet('Product Forecast')
    product_headers = [
        'Product Code',
        'Product',
        'Category',
        'Historical Sold',
        'Average Daily Quantity',
        'Forecast Quantity',
        'Projected Sales',
        'Projected Cost',
        'Projected Profit',
    ]
    product_sheet.append(product_headers)

    for cell in product_sheet[1]:
        cell.fill = dark_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal='center')

    for row in product_rows:
        product_sheet.append([
            row['product'].product_code or 'Auto',
            row['product'].display_name(),
            row['product'].display_category(),
            float(row['quantity_sold']),
            float(row['average_daily_quantity']),
            float(row['forecast_quantity']),
            float(row['projected_sales']),
            float(row['projected_cost']),
            float(row['projected_profit']),
        ])

    # Sheet 3: Ingredient Restock
    ingredient_sheet = workbook.create_sheet('Ingredient Restock')
    ingredient_headers = [
        'Ingredient',
        'Unit',
        'Projected Used',
        'Buffer Quantity',
        'Recommended Quantity',
        'Current Stock',
        'Reorder Level',
        'Suggested Purchase',
        'Status',
    ]
    ingredient_sheet.append(ingredient_headers)

    for cell in ingredient_sheet[1]:
        cell.fill = dark_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal='center')

    for row in ingredient_rows:
        ingredient_sheet.append([
            row['ingredient'].name,
            row['ingredient'].unit,
            float(row['projected_used_quantity']),
            float(row['buffer_quantity']),
            float(row['recommended_quantity']),
            float(row['current_quantity']),
            float(row['reorder_level']),
            float(row['suggested_purchase_quantity']),
            row['status'],
        ])

    # Sheet 4: Recipe Demand Details
    detail_sheet = workbook.create_sheet('Recipe Demand Details')
    detail_headers = [
        'Product Code',
        'Product',
        'Category',
        'Ingredient',
        'Forecast Quantity',
        'Required Per Product',
        'Projected Used',
        'Safety Buffer Percent',
        'Buffer Quantity',
        'Recommended Quantity',
        'Unit',
    ]
    detail_sheet.append(detail_headers)

    for cell in detail_sheet[1]:
        cell.fill = dark_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal='center')

    for row in detail_rows:
        detail_sheet.append([
            row['product'].product_code or 'Auto',
            row['product'].display_name(),
            row['product'].display_category(),
            row['ingredient'].name,
            float(row['forecast_quantity']),
            float(row['quantity_required']),
            float(row['projected_used_quantity']),
            float(row['safety_buffer_percent']),
            float(row['buffer_quantity']),
            float(row['recommended_quantity']),
            row['ingredient'].unit,
        ])

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.row == 1 or cell.value in ['Django Coffee System', 'Demand Forecasting Report']:
                    cell.font = bold_font

        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter

            for cell in column_cells:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))

            worksheet.column_dimensions[column_letter].width = max_length + 3

    filename = f"demand_forecasting_{date_from}_to_{date_to}_next_{forecast_days}_days.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    workbook.save(response)

    return response

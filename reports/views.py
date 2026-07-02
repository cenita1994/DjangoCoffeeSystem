import json
from decimal import Decimal
from datetime import datetime, timedelta

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Sum, Count, F, DecimalField, ExpressionWrapper, Q
from django.db.models.functions import TruncDate, TruncWeek, TruncMonth
from django.utils import timezone
from django.core.paginator import Paginator
from django.http import HttpResponse
from audittrail.utils import log_audit

from orders.models import Order, OrderItem
from inventory.models import Stock, ProductRecipeItem
from payments.models import WalletTransaction

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .models import CashierShift
from .forms import (
    CashierShiftOpenForm,
    CashierShiftCloseForm,
    CashierShiftReceiveForm,
    CashierShiftFilterForm,
)

from accounts.decorators import (
    manager_or_owner_required,
    cashier_manager_or_owner_required,
)


def decimal_to_float(value):
    if value is None:
        return 0

    return float(value)


def integer_value(value):
    if value is None:
        return 0

    return int(value)


def decimal_value(value):
    if value is None:
        return Decimal('0.00')

    return value


def is_manager_or_owner(user):
    return user.groups.filter(
        name__in=['Manager', 'Owner']
    ).exists()


def user_can_access_shift(user, shift):
    if is_manager_or_owner(user):
        return True

    return shift.cashier == user


def get_active_shift(cashier):
    return CashierShift.objects.filter(
        cashier=cashier,
        status='Open'
    ).first()


def calculate_shift_totals(shift):
    start_time = shift.shift_start

    if shift.shift_end:
        end_time = shift.shift_end
    else:
        end_time = timezone.now()

    paid_orders = Order.objects.filter(
        payment_status='Paid',
        payment_received_by=shift.cashier,
        paid_at__gte=start_time,
        paid_at__lte=end_time
    ).exclude(
        status='Cancelled'
    )

    cash_sales = paid_orders.filter(
        payment_method='Cash'
    ).aggregate(
        total=Sum('total_amount')
    )['total'] or Decimal('0.00')

    gcash_sales = paid_orders.filter(
        payment_method='GCash'
    ).aggregate(
        total=Sum('total_amount')
    )['total'] or Decimal('0.00')

    wallet_sales = paid_orders.filter(
        payment_method='Wallet'
    ).aggregate(
        total=Sum('total_amount')
    )['total'] or Decimal('0.00')

    total_sales_processed = paid_orders.aggregate(
        total=Sum('total_amount')
    )['total'] or Decimal('0.00')

    cash_in_received = WalletTransaction.objects.filter(
        transaction_type='Cash In',
        payment_method='Cash',
        performed_by=shift.cashier,
        transaction_date__gte=start_time,
        transaction_date__lte=end_time
    ).aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')

    gcash_cash_in = WalletTransaction.objects.filter(
        transaction_type='Cash In',
        payment_method='GCash',
        performed_by=shift.cashier,
        transaction_date__gte=start_time,
        transaction_date__lte=end_time
    ).aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')

    cash_refunds = WalletTransaction.objects.filter(
        transaction_type='Refund',
        payment_method='Cash',
        performed_by=shift.cashier,
        transaction_date__gte=start_time,
        transaction_date__lte=end_time
    ).aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')

    expected_cash = (
        decimal_value(shift.opening_cash)
        + decimal_value(cash_sales)
        + decimal_value(cash_in_received)
        - decimal_value(cash_refunds)
        - decimal_value(shift.cash_payouts)
    )

    if shift.actual_cash_counted is not None:
        over_short_amount = decimal_value(shift.actual_cash_counted) - expected_cash
    else:
        over_short_amount = Decimal('0.00')

    return {
        'cash_sales': cash_sales,
        'gcash_sales': gcash_sales,
        'wallet_sales': wallet_sales,
        'cash_in_received': cash_in_received,
        'gcash_cash_in': gcash_cash_in,
        'cash_refunds': cash_refunds,
        'expected_cash': expected_cash,
        'over_short_amount': over_short_amount,
        'total_sales_processed': total_sales_processed,
        'paid_orders_count': paid_orders.count(),
    }


def update_shift_computed_fields(shift):
    totals = calculate_shift_totals(shift)

    shift.cash_sales = totals['cash_sales']
    shift.gcash_sales = totals['gcash_sales']
    shift.wallet_sales = totals['wallet_sales']
    shift.cash_in_received = totals['cash_in_received']
    shift.gcash_cash_in = totals['gcash_cash_in']
    shift.cash_refunds = totals['cash_refunds']
    shift.expected_cash = totals['expected_cash']
    shift.over_short_amount = totals['over_short_amount']
    shift.total_sales_processed = totals['total_sales_processed']

    return shift

def parse_report_date(value, fallback):
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except (TypeError, ValueError):
        return fallback


def get_sales_period_filters(request):
    today = timezone.localdate()
    default_date_from = today.replace(day=1)

    date_from = parse_report_date(
        request.GET.get('date_from'),
        default_date_from
    )

    date_to = parse_report_date(
        request.GET.get('date_to'),
        today
    )

    if date_to < date_from:
        date_from, date_to = date_to, date_from

    group_by = request.GET.get('group_by', 'daily')

    if group_by not in ['daily', 'weekly', 'monthly']:
        group_by = 'daily'

    return date_from, date_to, group_by


def get_period_trunc(group_by, field_name):
    if group_by == 'weekly':
        return TruncWeek(field_name)

    if group_by == 'monthly':
        return TruncMonth(field_name)

    return TruncDate(field_name)


def normalize_period_date(period_value):
    if hasattr(period_value, 'date'):
        return period_value.date()

    return period_value


def get_period_label(period_start, group_by):
    if group_by == 'weekly':
        period_end = period_start + timedelta(days=6)
        return f"{period_start.strftime('%b %d, %Y')} - {period_end.strftime('%b %d, %Y')}"

    if group_by == 'monthly':
        return period_start.strftime('%B %Y')

    return period_start.strftime('%b %d, %Y')


def build_sales_period_rows(date_from, date_to, group_by):
    sales_orders = Order.objects.filter(
        payment_status='Paid',
        paid_at__date__gte=date_from,
        paid_at__date__lte=date_to
    ).exclude(
        status='Cancelled'
    )

    order_period_trunc = get_period_trunc(group_by, 'paid_at')

    sales_queryset = sales_orders.annotate(
        period=order_period_trunc
    ).values(
        'period'
    ).annotate(
        order_count=Count('id'),
        gross_sales=Sum('subtotal_amount'),
        total_discounts=Sum('discount_amount'),
        net_sales=Sum('total_amount'),
        cash_sales=Sum('total_amount', filter=Q(payment_method='Cash')),
        gcash_sales=Sum('total_amount', filter=Q(payment_method='GCash')),
        wallet_sales=Sum('total_amount', filter=Q(payment_method='Wallet')),
    ).order_by('period')

    item_period_trunc = get_period_trunc(group_by, 'order__paid_at')

    items_queryset = OrderItem.objects.filter(
        order__payment_status='Paid',
        order__paid_at__date__gte=date_from,
        order__paid_at__date__lte=date_to
    ).exclude(
        order__status='Cancelled'
    ).annotate(
        period=item_period_trunc
    ).values(
        'period'
    ).annotate(
        items_sold=Sum('quantity')
    )

    items_map = {}

    for item in items_queryset:
        period_start = normalize_period_date(item['period'])
        items_map[period_start.isoformat()] = integer_value(item['items_sold'])

    shift_period_trunc = get_period_trunc(group_by, 'shift_start')

    cash_payout_queryset = CashierShift.objects.filter(
        status='Closed',
        shift_start__date__gte=date_from,
        shift_start__date__lte=date_to
    ).annotate(
        period=shift_period_trunc
    ).values(
        'period'
    ).annotate(
        cash_payouts=Sum('cash_payouts')
    )

    cash_payout_map = {}

    for item in cash_payout_queryset:
        period_start = normalize_period_date(item['period'])
        cash_payout_map[period_start.isoformat()] = decimal_value(item['cash_payouts'])

    rows = []

    for item in sales_queryset:
        period_start = normalize_period_date(item['period'])
        period_key = period_start.isoformat()

        gross_sales = decimal_value(item['gross_sales'])
        total_discounts = decimal_value(item['total_discounts'])
        net_sales = decimal_value(item['net_sales'])
        cash_sales = decimal_value(item['cash_sales'])
        gcash_sales = decimal_value(item['gcash_sales'])
        wallet_sales = decimal_value(item['wallet_sales'])
        cash_payouts = decimal_value(cash_payout_map.get(period_key))
        estimated_net_after_payouts = net_sales - cash_payouts

        rows.append({
            'period_start': period_start,
            'period_label': get_period_label(period_start, group_by),
            'order_count': integer_value(item['order_count']),
            'items_sold': integer_value(items_map.get(period_key)),
            'gross_sales': gross_sales,
            'total_discounts': total_discounts,
            'net_sales': net_sales,
            'cash_sales': cash_sales,
            'gcash_sales': gcash_sales,
            'wallet_sales': wallet_sales,
            'cash_payouts': cash_payouts,
            'estimated_net_after_payouts': estimated_net_after_payouts,
        })

    summary = {
        'order_count': sum(row['order_count'] for row in rows),
        'items_sold': sum(row['items_sold'] for row in rows),
        'gross_sales': sum(row['gross_sales'] for row in rows),
        'total_discounts': sum(row['total_discounts'] for row in rows),
        'net_sales': sum(row['net_sales'] for row in rows),
        'cash_sales': sum(row['cash_sales'] for row in rows),
        'gcash_sales': sum(row['gcash_sales'] for row in rows),
        'wallet_sales': sum(row['wallet_sales'] for row in rows),
        'cash_payouts': sum(row['cash_payouts'] for row in rows),
        'estimated_net_after_payouts': sum(row['estimated_net_after_payouts'] for row in rows),
    }

    return rows, summary


@manager_or_owner_required
def reports_center(request):
    return render(request, 'reports/reports_center.html')


@manager_or_owner_required
def sales_period_report(request):
    date_from, date_to, group_by = get_sales_period_filters(request)

    rows, summary = build_sales_period_rows(
        date_from=date_from,
        date_to=date_to,
        group_by=group_by
    )

    return render(request, 'reports/sales_period_report.html', {
        'date_from': date_from,
        'date_to': date_to,
        'group_by': group_by,
        'rows': rows,
        'summary': summary,
    })


@manager_or_owner_required
def export_sales_period_report(request):
    log_audit(
        request=request,
        action='Export',
        module='Reports',
        description=f'Exported Sales Period Report. Filters: {request.GET.urlencode() or "None"}.',
        object_type='Report',
        object_repr='Sales Period Report'
    )

    date_from, date_to, group_by = get_sales_period_filters(request)

    rows, summary = build_sales_period_rows(
        date_from=date_from,
        date_to=date_to,
        group_by=group_by
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Sales by Period'

    dark_fill = PatternFill(
        fill_type='solid',
        fgColor='212529'
    )

    white_font = Font(
        color='FFFFFF',
        bold=True
    )

    bold_font = Font(bold=True)

    worksheet.append(['Django Coffee System'])
    worksheet.append(['Sales by Period Report'])
    worksheet.append(['Date From', date_from.strftime('%Y-%m-%d')])
    worksheet.append(['Date To', date_to.strftime('%Y-%m-%d')])
    worksheet.append(['Group By', group_by.title()])
    worksheet.append([])

    headers = [
        'Period',
        'Order Count',
        'Items Sold',
        'Gross Sales',
        'Discounts',
        'Net Sales',
        'Cash Sales',
        'GCash Sales',
        'Wallet Sales',
        'Cash Payouts / Expenses',
        'Estimated Net After Payouts',
    ]

    worksheet.append(headers)

    header_row_number = worksheet.max_row

    for cell in worksheet[header_row_number]:
        cell.fill = dark_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal='center')

    for row in rows:
        worksheet.append([
            row['period_label'],
            row['order_count'],
            row['items_sold'],
            float(row['gross_sales']),
            float(row['total_discounts']),
            float(row['net_sales']),
            float(row['cash_sales']),
            float(row['gcash_sales']),
            float(row['wallet_sales']),
            float(row['cash_payouts']),
            float(row['estimated_net_after_payouts']),
        ])

    worksheet.append([])

    worksheet.append([
        'TOTAL',
        summary['order_count'],
        summary['items_sold'],
        float(summary['gross_sales']),
        float(summary['total_discounts']),
        float(summary['net_sales']),
        float(summary['cash_sales']),
        float(summary['gcash_sales']),
        float(summary['wallet_sales']),
        float(summary['cash_payouts']),
        float(summary['estimated_net_after_payouts']),
    ])

    total_row_number = worksheet.max_row

    for cell in worksheet[total_row_number]:
        cell.font = bold_font

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        worksheet.column_dimensions[column_letter].width = max_length + 3

    filename = f"sales_period_{date_from}_to_{date_to}_{group_by}.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    workbook.save(response)

    return response


def get_product_sales_filters(request):
    today = timezone.localdate()
    default_date_from = today.replace(day=1)

    date_from = parse_report_date(
        request.GET.get('date_from'),
        default_date_from
    )

    date_to = parse_report_date(
        request.GET.get('date_to'),
        today
    )

    if date_to < date_from:
        date_from, date_to = date_to, date_from

    category = request.GET.get('category', '').strip()

    return date_from, date_to, category


def get_product_categories():
    categories = OrderItem.objects.exclude(
        product__category__isnull=True
    ).exclude(
        product__category=''
    ).values_list(
        'product__category',
        flat=True
    ).distinct().order_by(
        'product__category'
    )

    return categories


def build_product_sales_rows(date_from, date_to, category):
    product_items = OrderItem.objects.filter(
        order__payment_status='Paid',
        order__paid_at__isnull=False,
        order__paid_at__date__gte=date_from,
        order__paid_at__date__lte=date_to
    ).exclude(
        order__status='Cancelled'
    ).select_related(
        'order',
        'product',
        'product__product_category'
    ).prefetch_related(
        'discounts'
    )

    if category:
        product_items = product_items.filter(
            product__category=category
        )

    product_map = {}

    for item in product_items:
        product = item.product

        product_key = (
            product.id,
            product.name,
            product.size,
            product.display_category()
        )

        if product_key not in product_map:
            product_map[product_key] = {
                'product_id': product.id,
                'product_code': product.product_code or '',
                'product_name': product.name or 'Unknown Product',
                'size': product.size or 'Not Applicable',
                'category': product.display_category(),
                'order_ids': set(),
                'quantity_sold': 0,
                'gross_sales': Decimal('0.00'),
                'total_discounts': Decimal('0.00'),
                'net_sales': Decimal('0.00'),
                'product_cost': Decimal('0.00'),
                'estimated_gross_profit': Decimal('0.00'),
                'average_selling_price': Decimal('0.00'),
            }

        quantity = integer_value(item.quantity)
        item_price = decimal_value(item.price)
        item_cost_price = decimal_value(item.cost_price)

        gross_sales = item_price * quantity
        product_cost = item_cost_price * quantity
        total_discounts = decimal_value(item.total_discount())
        net_sales = gross_sales - total_discounts

        if net_sales < 0:
            net_sales = Decimal('0.00')

        product_map[product_key]['order_ids'].add(item.order_id)
        product_map[product_key]['quantity_sold'] += quantity
        product_map[product_key]['gross_sales'] += gross_sales
        product_map[product_key]['total_discounts'] += total_discounts
        product_map[product_key]['net_sales'] += net_sales
        product_map[product_key]['product_cost'] += product_cost

    rows = []

    for data in product_map.values():
        quantity_sold = data['quantity_sold']

        if quantity_sold > 0:
            data['average_selling_price'] = data['gross_sales'] / quantity_sold
        else:
            data['average_selling_price'] = Decimal('0.00')

        data['order_count'] = len(data['order_ids'])
        data['estimated_gross_profit'] = data['net_sales'] - data['product_cost']

        del data['order_ids']

        rows.append(data)

    rows = sorted(
        rows,
        key=lambda row: row['net_sales'],
        reverse=True
    )

    summary = {
        'product_count': len(rows),
        'order_count': sum(row['order_count'] for row in rows),
        'quantity_sold': sum(row['quantity_sold'] for row in rows),
        'gross_sales': sum(row['gross_sales'] for row in rows),
        'total_discounts': sum(row['total_discounts'] for row in rows),
        'net_sales': sum(row['net_sales'] for row in rows),
        'product_cost': sum(row['product_cost'] for row in rows),
        'estimated_gross_profit': sum(row['estimated_gross_profit'] for row in rows),
    }

    return rows, summary


@manager_or_owner_required
def product_sales_report(request):
    date_from, date_to, category = get_product_sales_filters(request)

    rows, summary = build_product_sales_rows(
        date_from=date_from,
        date_to=date_to,
        category=category
    )

    categories = get_product_categories()

    return render(request, 'reports/product_sales_report.html', {
        'date_from': date_from,
        'date_to': date_to,
        'category': category,
        'categories': categories,
        'rows': rows,
        'summary': summary,
    })

@manager_or_owner_required
def export_product_sales_report(request):
    log_audit(
        request=request,
        action='Export',
        module='Reports',
        description=f'Exported Product Sales Report. Filters: {request.GET.urlencode() or "None"}.',
        object_type='Report',
        object_repr='Product Sales Report'
    )

    date_from, date_to, category = get_product_sales_filters(request)

    rows, summary = build_product_sales_rows(
        date_from=date_from,
        date_to=date_to,
        category=category
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Product Sales'

    dark_fill = PatternFill(
        fill_type='solid',
        fgColor='212529'
    )

    white_font = Font(
        color='FFFFFF',
        bold=True
    )

    bold_font = Font(bold=True)

    worksheet.append(['Django Coffee System'])
    worksheet.append(['Product Sales and Profit Report'])
    worksheet.append(['Date From', date_from.strftime('%Y-%m-%d')])
    worksheet.append(['Date To', date_to.strftime('%Y-%m-%d')])
    worksheet.append(['Category', category if category else 'All Categories'])
    worksheet.append([])
    worksheet.append([
        'Note',
        'Estimated gross profit is computed as Net Sales minus saved product cost at the time of sale.'
    ])
    worksheet.append([])

    headers = [
        'Product Code',
        'Product',
        'Size',
        'Category',
        'Order Count',
        'Quantity Sold',
        'Gross Sales',
        'Discounts',
        'Net Sales',
        'Product Cost',
        'Estimated Gross Profit',
        'Average Selling Price',
    ]

    worksheet.append(headers)

    header_row_number = worksheet.max_row

    for cell in worksheet[header_row_number]:
        cell.fill = dark_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal='center')

    for row in rows:
        worksheet.append([
            row['product_code'],
            row['product_name'],
            row['size'],
            row['category'],
            row['order_count'],
            row['quantity_sold'],
            float(row['gross_sales']),
            float(row['total_discounts']),
            float(row['net_sales']),
            float(row['product_cost']),
            float(row['estimated_gross_profit']),
            float(row['average_selling_price']),
        ])

    worksheet.append([])

    worksheet.append([
        'TOTAL',
        '',
        '',
        '',
        summary['order_count'],
        summary['quantity_sold'],
        float(summary['gross_sales']),
        float(summary['total_discounts']),
        float(summary['net_sales']),
        float(summary['product_cost']),
        float(summary['estimated_gross_profit']),
        '',
    ])

    total_row_number = worksheet.max_row

    for cell in worksheet[total_row_number]:
        cell.font = bold_font

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        worksheet.column_dimensions[column_letter].width = max_length + 3

    if category:
        category_label = category.replace(' ', '_')
    else:
        category_label = 'all_categories'

    filename = f"product_sales_profit_{date_from}_to_{date_to}_{category_label}.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    workbook.save(response)

    return response

def get_payment_method_filters(request):
    today = timezone.localdate()
    default_date_from = today.replace(day=1)

    date_from = parse_report_date(
        request.GET.get('date_from'),
        default_date_from
    )

    date_to = parse_report_date(
        request.GET.get('date_to'),
        today
    )

    if date_to < date_from:
        date_from, date_to = date_to, date_from

    return date_from, date_to


def build_payment_method_rows(date_from, date_to):
    sales_orders = Order.objects.filter(
        payment_status='Paid',
        paid_at__date__gte=date_from,
        paid_at__date__lte=date_to
    ).exclude(
        status='Cancelled'
    )

    payment_queryset = sales_orders.values(
        'payment_method'
    ).annotate(
        order_count=Count('id'),
        gross_sales=Sum('subtotal_amount'),
        total_discounts=Sum('discount_amount'),
        net_sales=Sum('total_amount'),
    ).order_by(
        'payment_method'
    )

    items_queryset = OrderItem.objects.filter(
        order__payment_status='Paid',
        order__paid_at__date__gte=date_from,
        order__paid_at__date__lte=date_to
    ).exclude(
        order__status='Cancelled'
    ).values(
        'order__payment_method'
    ).annotate(
        items_sold=Sum('quantity')
    )

    items_map = {}

    for item in items_queryset:
        payment_method = item['order__payment_method'] or 'Unknown'
        items_map[payment_method] = integer_value(item['items_sold'])

    payment_map = {}

    for item in payment_queryset:
        payment_method = item['payment_method'] or 'Unknown'

        order_count = integer_value(item['order_count'])
        items_sold = integer_value(items_map.get(payment_method))

        gross_sales = decimal_value(item['gross_sales'])
        total_discounts = decimal_value(item['total_discounts'])
        net_sales = decimal_value(item['net_sales'])

        if order_count > 0:
            average_order_value = net_sales / order_count
        else:
            average_order_value = Decimal('0.00')

        payment_map[payment_method] = {
            'payment_method': payment_method,
            'order_count': order_count,
            'items_sold': items_sold,
            'gross_sales': gross_sales,
            'total_discounts': total_discounts,
            'net_sales': net_sales,
            'average_order_value': average_order_value,
        }

    rows = []

    for method in ['Cash', 'GCash', 'Wallet']:
        if method in payment_map:
            rows.append(payment_map[method])
        else:
            rows.append({
                'payment_method': method,
                'order_count': 0,
                'items_sold': 0,
                'gross_sales': Decimal('0.00'),
                'total_discounts': Decimal('0.00'),
                'net_sales': Decimal('0.00'),
                'average_order_value': Decimal('0.00'),
            })

    for method, data in payment_map.items():
        if method not in ['Cash', 'GCash', 'Wallet']:
            rows.append(data)

    summary = {
        'order_count': sum(row['order_count'] for row in rows),
        'items_sold': sum(row['items_sold'] for row in rows),
        'gross_sales': sum(row['gross_sales'] for row in rows),
        'total_discounts': sum(row['total_discounts'] for row in rows),
        'net_sales': sum(row['net_sales'] for row in rows),
    }

    if summary['order_count'] > 0:
        summary['average_order_value'] = summary['net_sales'] / summary['order_count']
    else:
        summary['average_order_value'] = Decimal('0.00')

    return rows, summary



@manager_or_owner_required
def payment_method_report(request):
    date_from, date_to = get_payment_method_filters(request)

    rows, summary = build_payment_method_rows(
        date_from=date_from,
        date_to=date_to
    )

    return render(request, 'reports/payment_method_report.html', {
        'date_from': date_from,
        'date_to': date_to,
        'rows': rows,
        'summary': summary,
    })

@manager_or_owner_required
def export_payment_method_report(request):
    log_audit(
        request=request,
        action='Export',
        module='Reports',
        description=f'Exported Payment Method Report. Filters: {request.GET.urlencode() or "None"}.',
        object_type='Report',
        object_repr='Payment Method Report'
    )

    date_from, date_to = get_payment_method_filters(request)

    rows, summary = build_payment_method_rows(
        date_from=date_from,
        date_to=date_to
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Payment Methods'

    dark_fill = PatternFill(
        fill_type='solid',
        fgColor='212529'
    )

    white_font = Font(
        color='FFFFFF',
        bold=True
    )

    bold_font = Font(bold=True)

    worksheet.append(['Django Coffee System'])
    worksheet.append(['Payment Method Report'])
    worksheet.append(['Date From', date_from.strftime('%Y-%m-%d')])
    worksheet.append(['Date To', date_to.strftime('%Y-%m-%d')])
    worksheet.append([])

    headers = [
        'Payment Method',
        'Order Count',
        'Items Sold',
        'Gross Sales',
        'Discounts',
        'Net Sales',
        'Average Order Value',
    ]

    worksheet.append(headers)

    header_row_number = worksheet.max_row

    for cell in worksheet[header_row_number]:
        cell.fill = dark_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal='center')

    for row in rows:
        worksheet.append([
            row['payment_method'],
            row['order_count'],
            row['items_sold'],
            float(row['gross_sales']),
            float(row['total_discounts']),
            float(row['net_sales']),
            float(row['average_order_value']),
        ])

    worksheet.append([])

    worksheet.append([
        'TOTAL',
        summary['order_count'],
        summary['items_sold'],
        float(summary['gross_sales']),
        float(summary['total_discounts']),
        float(summary['net_sales']),
        float(summary['average_order_value']),
    ])

    total_row_number = worksheet.max_row

    for cell in worksheet[total_row_number]:
        cell.font = bold_font

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        worksheet.column_dimensions[column_letter].width = max_length + 3

    filename = f"payment_methods_{date_from}_to_{date_to}.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    workbook.save(response)

    return response


def get_estimated_profit_filters(request):
    today = timezone.localdate()
    default_date_from = today.replace(day=1)

    date_from = parse_report_date(
        request.GET.get('date_from'),
        default_date_from
    )

    date_to = parse_report_date(
        request.GET.get('date_to'),
        today
    )

    if date_to < date_from:
        date_from, date_to = date_to, date_from

    group_by = request.GET.get('group_by', 'daily')

    if group_by not in ['daily', 'weekly', 'monthly']:
        group_by = 'daily'

    return date_from, date_to, group_by


def build_estimated_profit_rows(date_from, date_to, group_by):
    item_period_trunc = get_period_trunc(group_by, 'order__paid_at')

    order_items = OrderItem.objects.filter(
        order__payment_status='Paid',
        order__paid_at__isnull=False,
        order__paid_at__date__gte=date_from,
        order__paid_at__date__lte=date_to
    ).exclude(
        order__status='Cancelled'
    ).select_related(
        'order',
        'product'
    ).prefetch_related(
        'discounts'
    ).annotate(
        period=item_period_trunc
    ).order_by(
        'period'
    )

    profit_map = {}

    for item in order_items:
        period_start = normalize_period_date(item.period)
        period_key = period_start.isoformat()

        if period_key not in profit_map:
            profit_map[period_key] = {
                'period_start': period_start,
                'period_label': get_period_label(period_start, group_by),
                'order_ids': set(),
                'items_sold': 0,
                'gross_sales': Decimal('0.00'),
                'total_discounts': Decimal('0.00'),
                'net_sales': Decimal('0.00'),
                'product_cost': Decimal('0.00'),
                'gross_profit': Decimal('0.00'),
                'cash_payouts': Decimal('0.00'),
                'estimated_net_profit': Decimal('0.00'),
                'gross_profit_margin': Decimal('0.00'),
            }

        quantity = integer_value(item.quantity)
        item_price = decimal_value(item.price)
        item_cost_price = decimal_value(item.cost_price)

        item_gross_sales = item_price * quantity
        item_discount = decimal_value(item.total_discount())
        item_net_sales = item_gross_sales - item_discount
        item_product_cost = item_cost_price * quantity

        if item_net_sales < 0:
            item_net_sales = Decimal('0.00')

        item_gross_profit = item_net_sales - item_product_cost

        profit_map[period_key]['order_ids'].add(item.order_id)
        profit_map[period_key]['items_sold'] += quantity
        profit_map[period_key]['gross_sales'] += item_gross_sales
        profit_map[period_key]['total_discounts'] += item_discount
        profit_map[period_key]['net_sales'] += item_net_sales
        profit_map[period_key]['product_cost'] += item_product_cost
        profit_map[period_key]['gross_profit'] += item_gross_profit

    shift_period_trunc = get_period_trunc(group_by, 'shift_start')

    cash_payout_queryset = CashierShift.objects.filter(
        status='Closed',
        shift_start__date__gte=date_from,
        shift_start__date__lte=date_to
    ).annotate(
        period=shift_period_trunc
    ).values(
        'period'
    ).annotate(
        cash_payouts=Sum('cash_payouts')
    )

    for item in cash_payout_queryset:
        period_start = normalize_period_date(item['period'])
        period_key = period_start.isoformat()

        if period_key not in profit_map:
            profit_map[period_key] = {
                'period_start': period_start,
                'period_label': get_period_label(period_start, group_by),
                'order_ids': set(),
                'items_sold': 0,
                'gross_sales': Decimal('0.00'),
                'total_discounts': Decimal('0.00'),
                'net_sales': Decimal('0.00'),
                'product_cost': Decimal('0.00'),
                'gross_profit': Decimal('0.00'),
                'cash_payouts': Decimal('0.00'),
                'estimated_net_profit': Decimal('0.00'),
                'gross_profit_margin': Decimal('0.00'),
            }

        profit_map[period_key]['cash_payouts'] = decimal_value(item['cash_payouts'])

    rows = []

    for data in profit_map.values():
        data['order_count'] = len(data['order_ids'])
        data['estimated_net_profit'] = data['gross_profit'] - data['cash_payouts']

        if data['net_sales'] > 0:
            data['gross_profit_margin'] = (data['gross_profit'] / data['net_sales']) * Decimal('100')
        else:
            data['gross_profit_margin'] = Decimal('0.00')

        del data['order_ids']

        rows.append(data)

    rows = sorted(
        rows,
        key=lambda row: row['period_start']
    )

    summary = {
        'order_count': sum(row['order_count'] for row in rows),
        'items_sold': sum(row['items_sold'] for row in rows),
        'gross_sales': sum(row['gross_sales'] for row in rows),
        'total_discounts': sum(row['total_discounts'] for row in rows),
        'net_sales': sum(row['net_sales'] for row in rows),
        'product_cost': sum(row['product_cost'] for row in rows),
        'gross_profit': sum(row['gross_profit'] for row in rows),
        'cash_payouts': sum(row['cash_payouts'] for row in rows),
        'estimated_net_profit': sum(row['estimated_net_profit'] for row in rows),
    }

    if summary['net_sales'] > 0:
        summary['gross_profit_margin'] = (summary['gross_profit'] / summary['net_sales']) * Decimal('100')
    else:
        summary['gross_profit_margin'] = Decimal('0.00')

    return rows, summary


@manager_or_owner_required
def estimated_profit_report(request):
    date_from, date_to, group_by = get_estimated_profit_filters(request)

    rows, summary = build_estimated_profit_rows(
        date_from=date_from,
        date_to=date_to,
        group_by=group_by
    )

    return render(request, 'reports/estimated_profit_report.html', {
        'date_from': date_from,
        'date_to': date_to,
        'group_by': group_by,
        'rows': rows,
        'summary': summary,
    })


@manager_or_owner_required
def export_estimated_profit_report(request):
    log_audit(
        request=request,
        action='Export',
        module='Reports',
        description=f'Exported Estimated Profit Report. Filters: {request.GET.urlencode() or "None"}.',
        object_type='Report',
        object_repr='Estimated Profit Report'
    )

    date_from, date_to, group_by = get_estimated_profit_filters(request)

    rows, summary = build_estimated_profit_rows(
        date_from=date_from,
        date_to=date_to,
        group_by=group_by
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Estimated Profit'

    dark_fill = PatternFill(
        fill_type='solid',
        fgColor='212529'
    )

    white_font = Font(
        color='FFFFFF',
        bold=True
    )

    bold_font = Font(bold=True)

    worksheet.append(['Django Coffee System'])
    worksheet.append(['Estimated Profit Report'])
    worksheet.append(['Date From', date_from.strftime('%Y-%m-%d')])
    worksheet.append(['Date To', date_to.strftime('%Y-%m-%d')])
    worksheet.append(['Group By', group_by.title()])
    worksheet.append([])
    worksheet.append([
        'Note',
        'Estimated net profit is computed as Gross Profit minus recorded cash payouts. It is not final accounting profit.'
    ])
    worksheet.append([])

    headers = [
        'Period',
        'Order Count',
        'Items Sold',
        'Gross Sales',
        'Discounts',
        'Net Sales',
        'Product Cost',
        'Gross Profit',
        'Gross Profit Margin',
        'Cash Payouts / Expenses',
        'Estimated Net Profit',
    ]

    worksheet.append(headers)

    header_row_number = worksheet.max_row

    for cell in worksheet[header_row_number]:
        cell.fill = dark_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal='center')

    for row in rows:
        worksheet.append([
            row['period_label'],
            row['order_count'],
            row['items_sold'],
            float(row['gross_sales']),
            float(row['total_discounts']),
            float(row['net_sales']),
            float(row['product_cost']),
            float(row['gross_profit']),
            float(row['gross_profit_margin']),
            float(row['cash_payouts']),
            float(row['estimated_net_profit']),
        ])

    worksheet.append([])

    worksheet.append([
        'TOTAL',
        summary['order_count'],
        summary['items_sold'],
        float(summary['gross_sales']),
        float(summary['total_discounts']),
        float(summary['net_sales']),
        float(summary['product_cost']),
        float(summary['gross_profit']),
        float(summary['gross_profit_margin']),
        float(summary['cash_payouts']),
        float(summary['estimated_net_profit']),
    ])

    total_row_number = worksheet.max_row

    for cell in worksheet[total_row_number]:
        cell.font = bold_font

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        worksheet.column_dimensions[column_letter].width = max_length + 3

    filename = f"estimated_profit_{date_from}_to_{date_to}_{group_by}.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    workbook.save(response)

    return response

def get_cashier_sales_filters(request):
    today = timezone.localdate()
    default_date_from = today.replace(day=1)

    date_from = parse_report_date(
        request.GET.get('date_from'),
        default_date_from
    )

    date_to = parse_report_date(
        request.GET.get('date_to'),
        today
    )

    if date_to < date_from:
        date_from, date_to = date_to, date_from

    return date_from, date_to


def get_cashier_display_name(user_id, first_name, last_name, username):
    full_name = f"{first_name or ''} {last_name or ''}".strip()

    if full_name:
        return full_name

    if username:
        return username

    if user_id:
        return f"User #{user_id}"

    return "Unknown Cashier"


def build_cashier_sales_rows(date_from, date_to):
    sales_orders = Order.objects.filter(
        payment_status='Paid',
        paid_at__date__gte=date_from,
        paid_at__date__lte=date_to
    ).exclude(
        status='Cancelled'
    )

    cashier_queryset = sales_orders.values(
        'payment_received_by',
        'payment_received_by__first_name',
        'payment_received_by__last_name',
        'payment_received_by__username',
    ).annotate(
        order_count=Count('id'),
        gross_sales=Sum('subtotal_amount'),
        total_discounts=Sum('discount_amount'),
        net_sales=Sum('total_amount'),
    ).order_by(
        'payment_received_by__first_name',
        'payment_received_by__last_name',
        'payment_received_by__username'
    )

    items_queryset = OrderItem.objects.filter(
        order__payment_status='Paid',
        order__paid_at__date__gte=date_from,
        order__paid_at__date__lte=date_to
    ).exclude(
        order__status='Cancelled'
    ).values(
        'order__payment_received_by'
    ).annotate(
        items_sold=Sum('quantity')
    )

    items_map = {}

    for item in items_queryset:
        cashier_id = item['order__payment_received_by']
        items_map[cashier_id] = integer_value(item['items_sold'])

    rows = []

    for item in cashier_queryset:
        cashier_id = item['payment_received_by']

        cashier_name = get_cashier_display_name(
            cashier_id,
            item['payment_received_by__first_name'],
            item['payment_received_by__last_name'],
            item['payment_received_by__username']
        )

        order_count = integer_value(item['order_count'])
        items_sold = integer_value(items_map.get(cashier_id))

        gross_sales = decimal_value(item['gross_sales'])
        total_discounts = decimal_value(item['total_discounts'])
        net_sales = decimal_value(item['net_sales'])

        if order_count > 0:
            average_order_value = net_sales / order_count
        else:
            average_order_value = Decimal('0.00')

        rows.append({
            'cashier_id': cashier_id,
            'cashier_name': cashier_name,
            'order_count': order_count,
            'items_sold': items_sold,
            'gross_sales': gross_sales,
            'total_discounts': total_discounts,
            'net_sales': net_sales,
            'average_order_value': average_order_value,
        })

    rows = sorted(
        rows,
        key=lambda row: row['net_sales'],
        reverse=True
    )

    summary = {
        'cashier_count': len(rows),
        'order_count': sum(row['order_count'] for row in rows),
        'items_sold': sum(row['items_sold'] for row in rows),
        'gross_sales': sum(row['gross_sales'] for row in rows),
        'total_discounts': sum(row['total_discounts'] for row in rows),
        'net_sales': sum(row['net_sales'] for row in rows),
    }

    if summary['order_count'] > 0:
        summary['average_order_value'] = summary['net_sales'] / summary['order_count']
    else:
        summary['average_order_value'] = Decimal('0.00')

    return rows, summary


@manager_or_owner_required
def cashier_sales_report(request):
    date_from, date_to = get_cashier_sales_filters(request)

    rows, summary = build_cashier_sales_rows(
        date_from=date_from,
        date_to=date_to
    )

    return render(request, 'reports/cashier_sales_report.html', {
        'date_from': date_from,
        'date_to': date_to,
        'rows': rows,
        'summary': summary,
    })


@manager_or_owner_required
def export_cashier_sales_report(request):
    log_audit(
        request=request,
        action='Export',
        module='Reports',
        description=f'Exported Cashier Sales Report. Filters: {request.GET.urlencode() or "None"}.',
        object_type='Report',
        object_repr='Cashier Sales Report'
    )

    date_from, date_to = get_cashier_sales_filters(request)

    rows, summary = build_cashier_sales_rows(
        date_from=date_from,
        date_to=date_to
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Cashier Sales'

    dark_fill = PatternFill(
        fill_type='solid',
        fgColor='212529'
    )

    white_font = Font(
        color='FFFFFF',
        bold=True
    )

    bold_font = Font(bold=True)

    worksheet.append(['Django Coffee System'])
    worksheet.append(['Cashier Sales Report'])
    worksheet.append(['Date From', date_from.strftime('%Y-%m-%d')])
    worksheet.append(['Date To', date_to.strftime('%Y-%m-%d')])
    worksheet.append([])

    headers = [
        'Cashier',
        'Order Count',
        'Items Sold',
        'Gross Sales',
        'Discounts',
        'Net Sales',
        'Average Order Value',
    ]

    worksheet.append(headers)

    header_row_number = worksheet.max_row

    for cell in worksheet[header_row_number]:
        cell.fill = dark_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal='center')

    for row in rows:
        worksheet.append([
            row['cashier_name'],
            row['order_count'],
            row['items_sold'],
            float(row['gross_sales']),
            float(row['total_discounts']),
            float(row['net_sales']),
            float(row['average_order_value']),
        ])

    worksheet.append([])

    worksheet.append([
        'TOTAL',
        summary['order_count'],
        summary['items_sold'],
        float(summary['gross_sales']),
        float(summary['total_discounts']),
        float(summary['net_sales']),
        float(summary['average_order_value']),
    ])

    total_row_number = worksheet.max_row

    for cell in worksheet[total_row_number]:
        cell.font = bold_font

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        worksheet.column_dimensions[column_letter].width = max_length + 3

    filename = f"cashier_sales_{date_from}_to_{date_to}.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    workbook.save(response)

    return response

@manager_or_owner_required
def sales_dashboard(request):
    sales_orders = Order.objects.filter(
        payment_status='Paid'
    ).exclude(
        status='Cancelled'
    )

    cancelled_orders = Order.objects.filter(status='Cancelled')

    gross_sales = sales_orders.aggregate(
        total=Sum('subtotal_amount')
    )['total'] or 0

    total_discounts = sales_orders.aggregate(
        total=Sum('discount_amount')
    )['total'] or 0

    total_sales = sales_orders.aggregate(
        total=Sum('total_amount')
    )['total'] or 0

    vatable_sales = sales_orders.aggregate(
        total=Sum('vatable_sales')
    )['total'] or 0

    vat_exempt_sales = sales_orders.aggregate(
        total=Sum('vat_exempt_sales')
    )['total'] or 0

    vat_amount = sales_orders.aggregate(
        total=Sum('vat_amount')
    )['total'] or 0

    cash_sales = sales_orders.filter(
        payment_method='Cash'
    ).aggregate(
        total=Sum('total_amount')
    )['total'] or 0

    gcash_sales = sales_orders.filter(
        payment_method='GCash'
    ).aggregate(
        total=Sum('total_amount')
    )['total'] or 0

    wallet_sales = sales_orders.filter(
        payment_method='Wallet'
    ).aggregate(
        total=Sum('total_amount')
    )['total'] or 0

    paid_order_count = sales_orders.count()
    cancelled_count = cancelled_orders.count()

    pending_count = Order.objects.filter(status='Pending').count()
    preparing_count = Order.objects.filter(status='Preparing').count()
    ready_count = Order.objects.filter(status='Ready').count()
    completed_count = Order.objects.filter(status='Completed').count()

    daily_sales_queryset = sales_orders.annotate(
        sales_date=TruncDate('paid_at')
    ).values(
        'sales_date'
    ).annotate(
        total_orders=Count('id'),
        total_sales=Sum('total_amount')
    ).order_by('-sales_date')[:7]

    daily_sales = list(reversed(list(daily_sales_queryset)))

    daily_sales_chart = []

    for item in daily_sales:
        sales_date = item['sales_date']

        if sales_date:
            date_label = sales_date.strftime('%b %d')
        else:
            date_label = 'No Date'

        daily_sales_chart.append({
            'date': date_label,
            'sales': decimal_to_float(item['total_sales']),
            'orders': integer_value(item['total_orders']),
        })

    payment_method_queryset = sales_orders.values(
        'payment_method'
    ).annotate(
        total_sales=Sum('total_amount'),
        total_orders=Count('id')
    ).order_by('payment_method')

    payment_method_chart = []

    for item in payment_method_queryset:
        payment_method_chart.append({
            'method': item['payment_method'] or 'Unknown',
            'sales': decimal_to_float(item['total_sales']),
            'orders': integer_value(item['total_orders']),
        })

    line_total = ExpressionWrapper(
        F('price') * F('quantity'),
        output_field=DecimalField(max_digits=12, decimal_places=2)
    )

    product_sales_queryset = OrderItem.objects.filter(
        order__payment_status='Paid'
    ).exclude(
        order__status='Cancelled'
    ).values(
        'product__name',
        'product__category'
    ).annotate(
        total_quantity=Sum('quantity'),
        gross_sales=Sum(line_total)
    ).order_by('-total_quantity')[:10]

    product_sales_chart = []

    for item in product_sales_queryset:
        product_sales_chart.append({
            'product': item['product__name'] or 'Unknown Product',
            'quantity': integer_value(item['total_quantity']),
            'sales': decimal_to_float(item['gross_sales']),
        })

    category_sales_queryset = OrderItem.objects.filter(
        order__payment_status='Paid'
    ).exclude(
        order__status='Cancelled'
    ).values(
        'product__category'
    ).annotate(
        total_quantity=Sum('quantity'),
        gross_sales=Sum(line_total)
    ).order_by('-gross_sales')

    category_sales_chart = []

    for item in category_sales_queryset:
        category_name = item['product__category'] or 'Uncategorized'

        category_sales_chart.append({
            'category': category_name,
            'quantity': integer_value(item['total_quantity']),
            'sales': decimal_to_float(item['gross_sales']),
        })

    cashier_sales_queryset = sales_orders.filter(
        payment_received_by__isnull=False
    ).values(
        'payment_received_by__username'
    ).annotate(
        total_orders=Count('id'),
        total_sales=Sum('total_amount'),
        cash_sales=Sum('total_amount', filter=Q(payment_method='Cash')),
        gcash_sales=Sum('total_amount', filter=Q(payment_method='GCash')),
        wallet_sales=Sum('total_amount', filter=Q(payment_method='Wallet')),
    ).order_by('-total_sales')

    cashier_sales_chart = []

    for item in cashier_sales_queryset:
        cashier_sales_chart.append({
            'cashier': item['payment_received_by__username'] or 'Unknown Cashier',
            'totalSales': decimal_to_float(item['total_sales']),
            'cashSales': decimal_to_float(item['cash_sales']),
            'gcashSales': decimal_to_float(item['gcash_sales']),
            'walletSales': decimal_to_float(item['wallet_sales']),
            'orders': integer_value(item['total_orders']),
        })

    low_stock_products = Stock.objects.select_related('product').filter(
        quantity__lte=F('reorder_level')
    ).order_by('quantity')[:10]

    return render(request, 'reports/sales_dashboard.html', {
        'gross_sales': gross_sales,
        'total_discounts': total_discounts,
        'total_sales': total_sales,
        'vatable_sales': vatable_sales,
        'vat_exempt_sales': vat_exempt_sales,
        'vat_amount': vat_amount,

        'cash_sales': cash_sales,
        'gcash_sales': gcash_sales,
        'wallet_sales': wallet_sales,

        'paid_order_count': paid_order_count,
        'completed_count': completed_count,
        'cancelled_count': cancelled_count,
        'pending_count': pending_count,
        'preparing_count': preparing_count,
        'ready_count': ready_count,

        'daily_sales': daily_sales,
        'payment_method_queryset': payment_method_queryset,
        'product_sales_queryset': product_sales_queryset,
        'category_sales_queryset': category_sales_queryset,
        'cashier_sales_queryset': cashier_sales_queryset,
        'low_stock_products': low_stock_products,

        'daily_sales_chart': json.dumps(daily_sales_chart),
        'payment_method_chart': json.dumps(payment_method_chart),
        'product_sales_chart': json.dumps(product_sales_chart),
        'category_sales_chart': json.dumps(category_sales_chart),
        'cashier_sales_chart': json.dumps(cashier_sales_chart),
    })


@cashier_manager_or_owner_required
def open_cashier_shift(request):
    active_shift = get_active_shift(request.user)

    if active_shift:
        messages.info(request, 'You already have an open cashier shift.')
        return redirect('cashier_shift_detail', shift_id=active_shift.id)

    if request.method == 'POST':
        form = CashierShiftOpenForm(request.POST)

        if form.is_valid():
            shift = form.save(commit=False)
            shift.cashier = request.user
            shift.status = 'Open'
            shift.save()

            messages.success(request, 'Cashier shift was opened successfully.')
            return redirect('cashier_shift_detail', shift_id=shift.id)

    else:
        form = CashierShiftOpenForm()

    return render(request, 'reports/open_cashier_shift.html', {
        'form': form,
    })


@cashier_manager_or_owner_required
def close_cashier_shift(request, shift_id):
    shift = get_object_or_404(CashierShift, id=shift_id)

    if not user_can_access_shift(request.user, shift):
        messages.error(request, 'You are not allowed to close this shift.')
        return redirect('cashier_shift_list')

    if shift.status == 'Closed':
        messages.info(request, 'This cashier shift is already closed.')
        return redirect('cashier_shift_detail', shift_id=shift.id)

    if request.method == 'POST':
        form = CashierShiftCloseForm(request.POST, instance=shift)

        if form.is_valid():
            shift = form.save(commit=False)
            shift.shift_end = timezone.now()
            shift.status = 'Closed'

            if is_manager_or_owner(request.user):
                shift.received_by_manager = request.user

            update_shift_computed_fields(shift)
            shift.save()

            messages.success(request, 'Cashier shift was closed successfully.')
            return redirect('cashier_shift_detail', shift_id=shift.id)

    else:
        update_shift_computed_fields(shift)

        form = CashierShiftCloseForm(instance=shift)

    totals = calculate_shift_totals(shift)

    return render(request, 'reports/close_cashier_shift.html', {
        'form': form,
        'shift': shift,
        'totals': totals,
    })



@manager_or_owner_required
def receive_cashier_shift(request, shift_id):
    shift = get_object_or_404(CashierShift, id=shift_id)

    if shift.status != 'Closed':
        messages.error(request, 'Only closed shifts can be received by the manager or owner.')
        return redirect('cashier_shift_detail', shift_id=shift.id)

    if shift.received_by_manager:
        messages.info(request, 'This shift has already been received.')
        return redirect('cashier_shift_detail', shift_id=shift.id)

    if request.method == 'POST':
        form = CashierShiftReceiveForm(request.POST, instance=shift)

        if form.is_valid():
            shift = form.save(commit=False)
            shift.received_by_manager = request.user
            shift.received_at = timezone.now()
            shift.save()

            messages.success(request, 'Cashier shift was successfully received and verified.')
            return redirect('cashier_shift_detail', shift_id=shift.id)

    else:
        form = CashierShiftReceiveForm(instance=shift)

    return render(request, 'reports/receive_cashier_shift.html', {
        'form': form,
        'shift': shift,
    })

@cashier_manager_or_owner_required
def cashier_shift_detail(request, shift_id):
    shift = get_object_or_404(
        CashierShift.objects.select_related(
            'cashier',
            'received_by_manager'
        ),
        id=shift_id
    )

    if not user_can_access_shift(request.user, shift):
        messages.error(request, 'You are not allowed to view this shift.')
        return redirect('dashboard')

    if shift.status == 'Open':
        update_shift_computed_fields(shift)

    totals = calculate_shift_totals(shift)

    cash_drawer_chart = [
        {
            'label': 'Opening Cash',
            'amount': decimal_to_float(shift.opening_cash),
        },
        {
            'label': 'Cash Sales',
            'amount': decimal_to_float(totals['cash_sales']),
        },
        {
            'label': 'Cash-in Received',
            'amount': decimal_to_float(totals['cash_in_received']),
        },
        {
            'label': 'Cash Refunds',
            'amount': decimal_to_float(totals['cash_refunds']),
        },
        {
            'label': 'Cash Payouts',
            'amount': decimal_to_float(shift.cash_payouts),
        },
        {
            'label': 'Expected Cash',
            'amount': decimal_to_float(totals['expected_cash']),
        },
    ]

    if shift.actual_cash_counted is not None:
        cash_drawer_chart.append({
            'label': 'Actual Cash',
            'amount': decimal_to_float(shift.actual_cash_counted),
        })
    
    return render(request, 'reports/cashier_shift_detail.html', {
    'shift': shift,
    'totals': totals,
    'cash_drawer_chart': json.dumps(cash_drawer_chart),
    'is_manager_or_owner_user': is_manager_or_owner(request.user),
    })


@cashier_manager_or_owner_required
def cashier_shift_list(request):
    is_manager_or_owner_user = is_manager_or_owner(request.user)

    form = CashierShiftFilterForm(
        request.GET or None,
        user=request.user
    )

    shifts = CashierShift.objects.select_related(
        'cashier',
        'received_by_manager'
    ).all()

    # Cashier can only see own shifts.
    # Manager and Owner can see all shifts.
    if not is_manager_or_owner_user:
        shifts = shifts.filter(cashier=request.user)

    if form.is_valid():
        status = form.cleaned_data.get('status')
        date_from = form.cleaned_data.get('date_from')
        date_to = form.cleaned_data.get('date_to')

        if is_manager_or_owner_user:
            cashier = form.cleaned_data.get('cashier')

            if cashier:
                shifts = shifts.filter(cashier=cashier)

        if status:
            shifts = shifts.filter(status=status)

        if date_from:
            shifts = shifts.filter(shift_start__date__gte=date_from)

        if date_to:
            shifts = shifts.filter(shift_start__date__lte=date_to)

    paginator = Paginator(shifts, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    open_shift = get_active_shift(request.user)

    return render(request, 'reports/cashier_shift_list.html', {
        'form': form,
        'page_obj': page_obj,
        'open_shift': open_shift,
        'is_manager_or_owner_user': is_manager_or_owner_user,
    })
    
    
def get_default_previous_week_range():
    today = timezone.localdate()

    current_week_monday = today - timedelta(days=today.weekday())
    previous_monday = current_week_monday - timedelta(days=7)
    previous_sunday = current_week_monday - timedelta(days=1)

    return previous_monday, previous_sunday


def get_ingredient_demand_filters(request):
    default_date_from, default_date_to = get_default_previous_week_range()

    date_from = parse_report_date(
        request.GET.get('date_from'),
        default_date_from
    )

    date_to = parse_report_date(
        request.GET.get('date_to'),
        default_date_to
    )

    if date_to < date_from:
        date_from, date_to = date_to, date_from

    return date_from, date_to


def build_ingredient_demand_rows(date_from, date_to):
    order_items = OrderItem.objects.filter(
        order__payment_status='Paid',
        order__paid_at__isnull=False,
        order__paid_at__date__gte=date_from,
        order__paid_at__date__lte=date_to
    ).exclude(
        order__status='Cancelled'
    ).select_related(
        'order',
        'product',
        'product__product_category'
    )

    product_sales_map = {}

    for item in order_items:
        product_id = item.product_id

        if product_id not in product_sales_map:
            product_sales_map[product_id] = {
                'product': item.product,
                'quantity_sold': 0,
            }

        product_sales_map[product_id]['quantity_sold'] += integer_value(item.quantity)

    recipe_items = ProductRecipeItem.objects.filter(
        is_active=True,
        product_id__in=product_sales_map.keys(),
        ingredient__is_active=True
    ).select_related(
        'product',
        'product__product_category',
        'ingredient'
    ).order_by(
        'ingredient__name',
        'product__name',
        'product__size'
    )

    detail_rows = []
    ingredient_summary_map = {}

    for recipe_item in recipe_items:
        product_data = product_sales_map.get(recipe_item.product_id)

        if not product_data:
            continue

        product = product_data['product']
        ingredient = recipe_item.ingredient

        quantity_sold = integer_value(product_data['quantity_sold'])
        quantity_required = decimal_value(recipe_item.quantity_required)
        safety_buffer_percent = decimal_value(ingredient.safety_buffer_percent)

        estimated_used_quantity = Decimal(str(quantity_sold)) * quantity_required
        buffer_quantity = estimated_used_quantity * safety_buffer_percent / Decimal('100')
        recommended_quantity = estimated_used_quantity + buffer_quantity

        detail_rows.append({
            'product': product,
            'ingredient': ingredient,
            'quantity_sold': quantity_sold,
            'quantity_required': quantity_required,
            'estimated_used_quantity': estimated_used_quantity,
            'safety_buffer_percent': safety_buffer_percent,
            'buffer_quantity': buffer_quantity,
            'recommended_quantity': recommended_quantity,
        })

        ingredient_id = ingredient.id

        if ingredient_id not in ingredient_summary_map:
            ingredient_summary_map[ingredient_id] = {
                'ingredient': ingredient,
                'estimated_used_quantity': Decimal('0.000'),
                'buffer_quantity': Decimal('0.000'),
                'recommended_quantity': Decimal('0.000'),
                'current_quantity': decimal_value(ingredient.current_quantity),
                'suggested_purchase_quantity': Decimal('0.000'),
            }

        ingredient_summary_map[ingredient_id]['estimated_used_quantity'] += estimated_used_quantity
        ingredient_summary_map[ingredient_id]['buffer_quantity'] += buffer_quantity
        ingredient_summary_map[ingredient_id]['recommended_quantity'] += recommended_quantity

    summary_rows = []

    for data in ingredient_summary_map.values():
        suggested_purchase_quantity = data['recommended_quantity'] - data['current_quantity']

        if suggested_purchase_quantity < 0:
            suggested_purchase_quantity = Decimal('0.000')

        data['suggested_purchase_quantity'] = suggested_purchase_quantity
        summary_rows.append(data)

    summary_rows = sorted(
        summary_rows,
        key=lambda row: row['ingredient'].name
    )

    summary = {
        'product_count': len(product_sales_map),
        'recipe_item_count': len(detail_rows),
        'ingredient_count': len(summary_rows),
        'total_suggested_purchase_items': sum(
            1 for row in summary_rows if row['suggested_purchase_quantity'] > 0
        ),
    }

    return detail_rows, summary_rows, summary


@manager_or_owner_required
def ingredient_demand_report(request):
    date_from, date_to = get_ingredient_demand_filters(request)

    detail_rows, summary_rows, summary = build_ingredient_demand_rows(
        date_from=date_from,
        date_to=date_to
    )

    return render(request, 'reports/ingredient_demand_report.html', {
        'date_from': date_from,
        'date_to': date_to,
        'detail_rows': detail_rows,
        'summary_rows': summary_rows,
        'summary': summary,
    })
    
    
@manager_or_owner_required
def export_ingredient_demand_report(request):
    log_audit(
        request=request,
        action='Export',
        module='Reports',
        description=f'Exported Ingredient Demand Report. Filters: {request.GET.urlencode() or "None"}.',
        object_type='Report',
        object_repr='Ingredient Demand Report'
    )

    date_from, date_to = get_ingredient_demand_filters(request)

    detail_rows, summary_rows, summary = build_ingredient_demand_rows(
        date_from=date_from,
        date_to=date_to
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Ingredient Demand'

    worksheet.append(['Django Coffee System'])
    worksheet.append(['Ingredient Demand Recommendation Report'])
    worksheet.append(['Date From', date_from.strftime('%Y-%m-%d')])
    worksheet.append(['Date To', date_to.strftime('%Y-%m-%d')])
    worksheet.append([])

    worksheet.append(['Ingredient Recommendation Summary'])

    dark_fill = PatternFill(
        fill_type='solid',
        fgColor='212529'
    )

    white_font = Font(
        color='FFFFFF',
        bold=True
    )

    summary_headers = [
        'Ingredient',
        'Unit',
        'Estimated Used',
        'Buffer Quantity',
        'Recommended Quantity',
        'Current Stock',
        'Suggested Purchase',
        'Status',
    ]

    worksheet.append(summary_headers)

    header_row_number = worksheet.max_row

    for cell in worksheet[header_row_number]:
        cell.fill = dark_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal='center')

    for row in summary_rows:
        status = 'Restock Suggested'

        if row['suggested_purchase_quantity'] <= 0:
            status = 'Enough Stock'

        worksheet.append([
            row['ingredient'].name,
            row['ingredient'].unit,
            float(row['estimated_used_quantity']),
            float(row['buffer_quantity']),
            float(row['recommended_quantity']),
            float(row['current_quantity']),
            float(row['suggested_purchase_quantity']),
            status,
        ])

    worksheet.append([])
    worksheet.append(['Product Recipe Demand Details'])

    detail_headers = [
        'Product Code',
        'Product / Variant',
        'Category',
        'Ingredient',
        'Quantity Sold',
        'Required Per Product',
        'Estimated Used',
        'Buffer Percent',
        'Buffer Quantity',
        'Recommended Quantity',
        'Unit',
    ]

    worksheet.append(detail_headers)

    detail_header_row_number = worksheet.max_row

    for cell in worksheet[detail_header_row_number]:
        cell.fill = dark_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal='center')

    for row in detail_rows:
        worksheet.append([
            row['product'].product_code or 'Auto',
            row['product'].display_name(),
            row['product'].display_category(),
            row['ingredient'].name,
            row['quantity_sold'],
            float(row['quantity_required']),
            float(row['estimated_used_quantity']),
            float(row['safety_buffer_percent']),
            float(row['buffer_quantity']),
            float(row['recommended_quantity']),
            row['ingredient'].unit,
        ])

    
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        worksheet.column_dimensions[column_letter].width = max_length + 3

    filename = f"ingredient_demand_{date_from}_to_{date_to}.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    workbook.save(response)

    return response